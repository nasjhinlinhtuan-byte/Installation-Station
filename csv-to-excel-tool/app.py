from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    csv_file = request.files["csv_file"]
    excel_template = request.files["excel_template"]

    # Đọc CSV
    df = pd.read_csv(csv_file)

    # Mở Excel mẫu
    wb = load_workbook(excel_template)
    ws = wb.active  # hoặc wb["TênSheet"]

    # Ví dụ: điền dòng đầu tiên của CSV vào Excel
    ws["B2"] = df["name"][0]
    ws["B3"] = df["email"][0]
    ws["B4"] = df["message"][0]

    # Lưu file tạm
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)

    return send_file(temp.name, as_attachment=True, download_name="result.xlsx")

if __name__ == "__main__":
    app.run()
